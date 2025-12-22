# apps/lms/utils/export.py
from __future__ import annotations

import csv
import io
from datetime import datetime
from typing import Iterable, List, Dict

from django.http import HttpResponse
from django.utils.timezone import now

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except Exception:  # pragma: no cover
    openpyxl = None

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
except Exception:  # pragma: no cover
    canvas = None


def _safe_filename(prefix: str) -> str:
    ts = now().strftime("%Y%m%d-%H%M%S")
    return f"{prefix}-{ts}"


def export_csv(filename_prefix: str, headers: List[str], rows: Iterable[Iterable]):
    """
    Trả về HttpResponse CSV.
    """
    filename = _safe_filename(filename_prefix) + ".csv"
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    for r in rows:
        writer.writerow(r)
    resp = HttpResponse(buf.getvalue(), content_type="text/csv; charset=utf-8")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def export_xlsx(filename_prefix: str, headers: List[str], rows: Iterable[Iterable]):
    """
    Trả về HttpResponse XLSX.
    """
    if openpyxl is None:
        # Fallback sang CSV nếu thiếu openpyxl
        return export_csv(filename_prefix, headers, rows)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    # header
    for ci, h in enumerate(headers, start=1):
        ws.cell(row=1, column=ci, value=h)

    # rows
    for ri, r in enumerate(rows, start=2):
        for ci, val in enumerate(r, start=1):
            ws.cell(row=ri, column=ci, value=val)

    # autosize
    for ci in range(1, len(headers) + 1):
        col_letter = get_column_letter(ci)
        ws.column_dimensions[col_letter].auto_size = True

    filename = _safe_filename(filename_prefix) + ".xlsx"
    buf = io.BytesIO()
    wb.save(buf)
    resp = HttpResponse(
        buf.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def export_pdf(filename_prefix: str, title: str, headers: List[str], rows: Iterable[Iterable]):
    """
    Trả về HttpResponse PDF đơn giản. Nếu thiếu reportlab => fallback CSV.
    """
    if canvas is None:
        return export_csv(filename_prefix, headers, rows)

    filename = _safe_filename(filename_prefix) + ".pdf"
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)

    width, height = A4
    margin = 40
    y = height - margin

    c.setTitle(title)
    c.setFont("Helvetica-Bold", 14)
    c.drawString(margin, y, title)
    y -= 24

    # headers
    c.setFont("Helvetica-Bold", 10)
    header_line = " | ".join(headers)
    c.drawString(margin, y, header_line)
    y -= 16
    c.setFont("Helvetica", 10)

    # rows (simple wrap)
    for r in rows:
        line = " | ".join([str(x) if x is not None else "" for x in r])
        if y < margin + 20:
            c.showPage()
            y = height - margin
        c.drawString(margin, y, line)
        y -= 14

    c.showPage()
    c.save()
    resp = HttpResponse(buf.getvalue(), content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp
